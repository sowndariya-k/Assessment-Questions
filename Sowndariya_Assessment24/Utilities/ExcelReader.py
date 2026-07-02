import openpyxl


def get_data(file_path, sheet_name):

    workbook = openpyxl.load_workbook(file_path)

    sheet = workbook[sheet_name]

    data = []

    for row in range(2, sheet.max_row + 1):

        row_data = []

        for col in range(1, sheet.max_column + 1):

            row_data.append(
                sheet.cell(row, col).value
            )

        data.append(row_data)

    workbook.close()

    return data